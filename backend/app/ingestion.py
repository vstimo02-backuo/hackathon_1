from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".csv", ".json", ".xlsx"}


@dataclass(frozen=True)
class ParsedFile:
    filename: str
    format: str
    row_count: int
    fields: list[dict[str, Any]]
    records: list[dict[str, Any]]
    errors: list[str]

    @property
    def status(self) -> str:
        return "valid" if not self.errors else "invalid"

    def to_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "format": self.format,
            "status": self.status,
            "row_count": self.row_count,
            "fields": self.fields,
            "records": self.records,
            "errors": self.errors,
        }


def parse_file(filename: str, content: bytes) -> ParsedFile:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        return _invalid(filename, extension.removeprefix(".") or "unknown", ["Unsupported file format."])

    try:
        if extension == ".csv":
            rows = _parse_csv(content)
        elif extension == ".json":
            rows = _parse_json(content)
        else:
            rows = _parse_xlsx(content)
    except ValueError as error:
        return _invalid(filename, extension.removeprefix("."), [str(error)])

    return ParsedFile(
        filename=filename,
        format=extension.removeprefix("."),
        row_count=len(rows),
        fields=_extract_fields(rows),
        records=[_json_safe_record(row) for row in rows],
        errors=[],
    )


def _invalid(filename: str, file_format: str, errors: list[str]) -> ParsedFile:
    return ParsedFile(filename=filename, format=file_format, row_count=0, fields=[], records=[], errors=errors)


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = _decode_text(content)
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV file must include a header row.")
    return [{key: value for key, value in row.items()} for row in reader]


def _parse_json(content: bytes) -> list[dict[str, Any]]:
    text = _decode_text(content)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as error:
        raise ValueError(f"Malformed JSON: {error.msg}.") from error

    if isinstance(payload, list) and all(isinstance(item, dict) for item in payload):
        return payload
    if isinstance(payload, dict):
        if all(isinstance(value, list) for value in payload.values()):
            max_length = max((len(value) for value in payload.values()), default=0)
            return [
                {key: values[index] if index < len(values) else None for key, values in payload.items()}
                for index in range(max_length)
            ]
        return [payload]
    raise ValueError("JSON file must contain an object, an array of objects, or an object of arrays.")


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise ValueError("Malformed XLSX workbook.") from error

    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows:
        raise ValueError("XLSX file must include a header row.")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        raise ValueError("XLSX header row must include at least one field name.")
    return [
        {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        for row in rows[1:]
    ]


def _decode_text(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError("File must be UTF-8 encoded.") from error


def _extract_fields(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    field_names: list[str] = []
    for row in rows:
        for field_name in row:
            if field_name not in field_names:
                field_names.append(field_name)
    return [_field_metadata(field_name, [row.get(field_name) for row in rows]) for field_name in field_names]


def _field_metadata(field_name: str, values: list[Any]) -> dict[str, Any]:
    non_empty_values = [value for value in values if value not in (None, "")]
    total_count = len(values)
    fill_rate = round((len(non_empty_values) / total_count * 100), 2) if total_count > 0 else 0.0

    # PII / Sensitive data check heuristics
    pii_level = "None"
    name_lower = field_name.lower()
    if any(k in name_lower for k in ["tax", "vat", "ein", "tin", "ssn", "passport", "salary", "revenue"]):
        pii_level = "High"
    elif any(k in name_lower for k in ["mail", "phone", "tel", "mobile", "address", "loc", "street", "zip", "postcode", "contact"]):
        pii_level = "Medium"
    elif any(k in name_lower for k in ["id", "code", "uid", "name", "client", "customer"]):
        pii_level = "Low"

    return {
        "name": field_name,
        "inferred_type": _infer_type(non_empty_values),
        "non_empty_count": len(non_empty_values),
        "fill_rate": fill_rate,
        "pii_level": pii_level,
        "sample_values": [str(value) for value in non_empty_values[:3]],
    }


def _json_safe_record(row: dict[str, Any]) -> dict[str, Any]:
    return {key: _json_safe_value(value) for key, value in row.items()}


def _json_safe_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _infer_type(values: list[Any]) -> str:
    if not values:
        return "empty"
    detected = {_detect_type(value) for value in values}
    detected.discard("empty")
    return detected.pop() if len(detected) == 1 else "mixed"


def _detect_type(value: Any) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int) and not isinstance(value, bool):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, (date, datetime)):
        return "date"
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return "empty"
        if stripped.lower() in {"true", "false"}:
            return "boolean"
        try:
            int(stripped)
            return "integer"
        except ValueError:
            pass
        try:
            float(stripped)
            return "number"
        except ValueError:
            return "string"
    return "string"